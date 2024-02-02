from django.contrib import admin
from django import forms
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from django.http import HttpResponse
from django.utils import timezone
from .models import Materiais, Pacote, ItensPacote, SaidaMaterial, EntradaMaterial, Responsavel
import datetime


class ResponsavelAdmin(admin.ModelAdmin):
    list_display = ['nome', 'is_ativo', 'data_ativacao', 'data_encerramento']


class ItensPacoteInline(admin.TabularInline):
    model = ItensPacote


class PacoteAdmin(admin.ModelAdmin):
    inlines = [ItensPacoteInline]


class MateriaisAdmin(admin.ModelAdmin):
    list_display = ['nome', 'quantidade_em_estoque', 'categoria']
    ordering = ['nome']
    list_filter = ['categoria']
    actions = ['export_to_excel']

    def export_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        centered_alignment = Alignment(horizontal='center', vertical='center')

        headers = ['Nome', 'Quantidade em Estoque', 'Categoria']
        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 40
            ws[f"{col_letter}1"] = header_title
            ws[f"{col_letter}1"].font = Font(bold=True)
            ws[f"{col_letter}1"].alignment = centered_alignment

        row_num = 2
        for material in queryset:
            ws.cell(row=row_num, column=1).value = material.nome
            ws.cell(row=row_num, column=2).value = material.quantidade_em_estoque
            ws.cell(row=row_num, column=3).value = material.categoria if material.categoria else ''

            for col_num in range(1, 4):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"].border = thin_border
                ws[f"{col_letter}{row_num}"].alignment = centered_alignment
            row_num += 1

        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        file_name = f"Inventario - Jazon_{current_date}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        wb.save(response)

        return response

    export_to_excel.short_description = "Exportar selecionados para Excel"


class EntradaMaterialAdminForm(forms.ModelForm):
    class Meta:
        model = EntradaMaterial
        fields = '__all__'


class EntradaMaterialAdmin(admin.ModelAdmin):
    list_display = ['material', 'quantidade', 'data_entrada', 'remetente', 'responsavel']
    form = EntradaMaterialAdminForm
    list_filter = ['material__categoria', 'data_entrada', 'remetente']
    actions = ['export_to_excel']

    def save_model(self, request, obj, form, change):
        material = obj.material
        quantidade_entrada = obj.quantidade

        material.quantidade_em_estoque += quantidade_entrada
        material.save()

        super().save_model(request, obj, form, change)

    def export_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        centered_alignment = Alignment(horizontal='center', vertical='center')

        headers = ['Material', 'Quantidade', 'Data de Entrada', 'Remetente']
        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 45
            ws[f"{col_letter}1"] = header_title
            ws[f"{col_letter}1"].font = Font(bold=True)
            ws[f"{col_letter}1"].alignment = centered_alignment

        row_num = 2
        for obj in queryset:
            material = obj.material

            for item in material.itenspacote_set.all():
                ws.cell(row=row_num, column=1).value = item.material.nome
                ws.cell(row=row_num, column=2).value = obj.quantidade
                ws.cell(row=row_num, column=3).value = timezone.localtime(obj.data_entrada).replace(
                    tzinfo=None) if obj.data_entrada else None
                ws.cell(row=row_num, column=4).value = obj.remetente

                for col_num in range(1, 5):
                    col_letter = get_column_letter(col_num)
                    ws[f"{col_letter}{row_num}"].border = thin_border
                    ws[f"{col_letter}{row_num}"].alignment = centered_alignment

                row_num += 1

        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        file_name = f"Registro de Entrada_{current_date}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        wb.save(response)

        return response

    export_to_excel.short_description = "Exportar selecionados para Excel"


class SaidaMaterialAdminForm(forms.ModelForm):
    class Meta:
        model = SaidaMaterial
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        pacote = cleaned_data.get('pacote')

        for item in pacote.itenspacote_set.all():
            material = item.material
            quantidade_saida = item.quantidade

            if material.quantidade_em_estoque < quantidade_saida:
                msg = f"Quantidade insuficiente em estoque para {material.nome}."
                self.add_error(None, msg)


# Exportar saida de materiais em Excel:

class SaidaMaterialAdmin(admin.ModelAdmin):
    list_display = ['pacote', 'data_saida', 'destino', 'servico', 'responsavel']
    form = SaidaMaterialAdminForm

    actions = ['export_to_excel']

    def save_model(self, request, obj, form, change):
        pacote = obj.pacote

        for item in pacote.itenspacote_set.all():
            material = item.material
            quantidade_saida = item.quantidade

            if material.quantidade_em_estoque >= quantidade_saida:
                material.quantidade_em_estoque -= quantidade_saida
                material.save()
            else:
                msg = f"Quantidade insuficiente em estoque para {material.nome}."
                form.add_error(None, msg)
                return

        super().save_model(request, obj, form, change)

    def export_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        centered_alignment = Alignment(horizontal='center', vertical='center')

        headers = ['Material', 'Quantidade', 'Pacote', 'Data de Saída', 'Destino', 'Tipo de Serviço']
        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 45
            ws[f"{col_letter}1"] = header_title
            ws[f"{col_letter}1"].font = Font(bold=True)
            ws[f"{col_letter}1"].alignment = centered_alignment

        row_num = 2
        for obj in queryset:
            pacote = obj.pacote

            for item in pacote.itenspacote_set.all():
                ws.cell(row=row_num, column=1).value = item.material.nome
                ws.cell(row=row_num, column=2).value = item.quantidade
                ws.cell(row=row_num, column=3).value = pacote.nome
                ws.cell(row=row_num, column=4).value = timezone.localtime(obj.data_saida).replace(
                    tzinfo=None) if obj.data_saida else None
                ws.cell(row=row_num, column=5).value = obj.destino
                ws.cell(row=row_num, column=6).value = obj.servico

                for col_num in range(1, 7):
                    col_letter = get_column_letter(col_num)
                    ws[f"{col_letter}{row_num}"].border = thin_border
                    ws[f"{col_letter}{row_num}"].alignment = centered_alignment

                row_num += 1

        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        file_name = f"Registro de Saída_{current_date}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        wb.save(response)

        return response

    export_to_excel.short_description = "Exportar selecionados para Excel"


# admin.site.register(Materiais, MateriaisAdmin)
# admin.site.register(Pacote, PacoteAdmin)
# admin.site.register(EntradaMaterial, EntradaMaterialAdmin)
# admin.site.register(SaidaMaterial, SaidaMaterialAdmin)
# admin.site.register(Responsavel, ResponsavelAdmin)
