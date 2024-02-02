from django.contrib import admin
from django import forms
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum
from .models import Materiais, Pacote, ItensPacote, SaidaMaterial, EntradaMaterial, UnidadeArm

#Versão



class UnidadeArmAdmin(admin.ModelAdmin):
    list_display = ['nome', 'responsavel', 'is_ativo', 'data_ativacao', 'data_encerramento']
    actions = ['export_to_excel']

    def export_to_excel(self, request, queryset):
        # Criar um livro de trabalho e adicionar uma planilha
        wb = Workbook()
        ws = wb.active

        # Configurar estilos
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        centered_alignment = Alignment(horizontal='center', vertical='center')

        headers = ['Material', 'Quantidade', 'Categoria']
        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 40
            ws[f"{col_letter}1"] = header_title
            ws[f"{col_letter}1"].font = Font(bold=True)
            ws[f"{col_letter}1"].alignment = centered_alignment

        row_num = 2
        # Filtrar entradas de materiais por destino (unidade de armazenamento)
        materiais_quantidades = EntradaMaterial.objects.filter(destino__in=queryset).values('material__nome',
                                                                                            'material__categoria').annotate(
            quantidade_total=Sum('quantidade'))

        for item in materiais_quantidades:
            ws.cell(row=row_num, column=1).value = item['material__nome']
            ws.cell(row=row_num, column=2).value = item['quantidade_total']
            ws.cell(row=row_num, column=3).value = item['material__categoria']

            # Aplicar estilos
            for col_num in range(1, 4):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"].border = thin_border
                ws[f"{col_letter}{row_num}"].alignment = centered_alignment

            row_num += 1

        # Configurar o nome do arquivo
        current_date = timezone.now().strftime("%Y-%m-%d")
        file_name = f"Materiais_{current_date}.xlsx"

        # Configurar a resposta para download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        wb.save(response)

        return response

    export_to_excel.short_description = 'Exportar Materiais e Quantidades'


class ItensPacoteInline(admin.TabularInline):
    model = ItensPacote


class PacoteAdmin(admin.ModelAdmin):
    inlines = [ItensPacoteInline]


class MateriaisAdmin(admin.ModelAdmin):
    list_display = ['nome', 'quantidade_em_estoque', 'categoria']
    ordering = ['nome']
    list_filter = ['categoria']


class EntradaMaterialAdminForm(forms.ModelForm):
    class Meta:
        model = EntradaMaterial
        fields = '__all__'


class EntradaMaterialAdmin(admin.ModelAdmin):
    list_display = ['material', 'quantidade', 'data_entrada', 'remetente', 'destino']
    form = EntradaMaterialAdminForm
    list_filter = ['material__categoria', 'data_entrada', 'remetente']

    def save_model(self, request, obj, form, change):
        material = obj.material
        quantidade_entrada = obj.quantidade
        unidade_destino = obj.destino

        material.quantidade_em_estoque += quantidade_entrada
        material.save()

        super().save_model(request, obj, form, change)


class SaidaMaterialAdminForm(forms.ModelForm):
    class Meta:
        model = SaidaMaterial
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        pacote = cleaned_data.get('pacote')

        for item in pacote.itenspacote_set.all():
            material = item.material
            quantidade_saida = item.quantidade

            if material.quantidade_em_estoque < quantidade_saida:
                msg = f"Quantidade insuficiente em estoque para {material.nome}."
                self.add_error(None, msg)


class SaidaMaterialAdmin(admin.ModelAdmin):
    list_display = ['pacote', 'data_saida', 'destino', 'servico', 'unidade_debito']
    form = SaidaMaterialAdminForm

    def save_model(self, request, obj, form, change):
        pacote = obj.pacote
        unidade_debito = obj.unidade_debito
        transacao_bem_sucedida = True

        for item in pacote.itenspacote_set.all():
            material = item.material
            quantidade_saida = item.quantidade

            if material.quantidade_em_estoque >= quantidade_saida:
                material.quantidade_em_estoque -= quantidade_saida
                material.save()

                material.unidade_debito = unidade_debito
                material.save()
            else:
                msg = f"Quantidade insuficiente em estoque para {material.nome}."
                form.add_error(None, msg)
                transacao_bem_sucedida = False
                break

        if transacao_bem_sucedida:
            super().save_model(request, obj, form, change)
        else:
            form.add_error(None, "A transação de saída não foi concluída com sucesso.")


admin.site.register(Materiais, MateriaisAdmin)
admin.site.register(Pacote, PacoteAdmin)
admin.site.register(EntradaMaterial, EntradaMaterialAdmin)
admin.site.register(SaidaMaterial, SaidaMaterialAdmin)
admin.site.register(UnidadeArm, UnidadeArmAdmin)
